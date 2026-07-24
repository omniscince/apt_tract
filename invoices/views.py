from django.db import models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.views import View
import logging
logger = logging.getLogger('invoices')
import logging
logger = logging.getLogger('invoices')
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.mail import EmailMessage
from django.conf import settings
import io
import threading
from .models import Invoice, InvoiceItem
from .forms import InvoiceForm, InvoiceItemFormSet
from .pdf import generate_invoice_pdf, generate_monthly_report_pdf
from customers.models import Customer
from cars.models import Car


def get_or_create_customer_and_car(form, invoice):
    customer_name = form.cleaned_data.get('customer_name', '').strip()
    car_make = form.cleaned_data.get('car_make', '').strip()
    car_model = form.cleaned_data.get('car_model', '').strip()
    car_vin = form.cleaned_data.get('car_vin', '').strip()
    car_stock = form.cleaned_data.get('car_stock', '').strip()

    customer = Customer.objects.filter(name__iexact=customer_name).first()
    if not customer:
        invoice.customer_name = customer_name
        invoice.customer = None
    else:
        invoice.customer = customer
        invoice.customer_name = ''

    if customer and car_vin:
        car = Car.objects.filter(customer=customer, vin__iexact=car_vin).first()
        if not car:
            car = Car.objects.create(
                customer=customer,
                make=car_make,
                model=car_model,
                vin=car_vin,
                stock_number=car_stock,
            )
        else:
            # Update car make/model if changed
            car.make = car_make
            car.model = car_model
            car.stock_number = car_stock
            car.save()
        invoice.car = car
        invoice.car_info = ''
    elif customer and car_make:
        car = Car.objects.filter(
            customer=customer,
            make__iexact=car_make,
            model__iexact=car_model
        ).first()
        if not car:
            car = Car.objects.create(
                customer=customer,
                make=car_make,
                model=car_model,
                vin=car_vin,
                stock_number=car_stock,
            )
        else:
            car.make = car_make
            car.model = car_model
            car.stock_number = car_stock
            car.save()
        invoice.car = car
        invoice.car_info = ''
    else:
        invoice.car = None
        parts = [car_make]
        if car_model:
            parts.append(car_model)
        if car_vin:
            parts.append(f'VIN: {car_vin}')
        if car_stock:
            parts.append(f'Stock#: {car_stock}')
        invoice.car_info = ' '.join(parts)

    return invoice


@method_decorator(login_required, name='dispatch')
class DashboardView(View):
    def get(self, request):
        from datetime import date, timedelta
        user = request.user
        if user.role == 'staff':
            invoices_qs = Invoice.objects.filter(
                models.Q(created_by=user) | models.Q(work_completed_by=user)
            ).distinct()
        else:
            invoices_qs = Invoice.objects.all()

        period = request.GET.get('period', 'month')
        customer_id = request.GET.get('customer')
        staff_id = request.GET.get('staff')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        today = date.today()

        if period == 'today':
            invoices_qs = invoices_qs.filter(invoice_date=today)
        elif period == 'week':
            week_start = today - timedelta(days=today.weekday())
            invoices_qs = invoices_qs.filter(invoice_date__gte=week_start, invoice_date__lte=today)
        elif period == 'month':
            invoices_qs = invoices_qs.filter(invoice_date__year=today.year, invoice_date__month=today.month)
        elif period == 'custom':
            if date_from:
                invoices_qs = invoices_qs.filter(invoice_date__gte=date_from)
            if date_to:
                invoices_qs = invoices_qs.filter(invoice_date__lte=date_to)

        if customer_id:
            invoices_qs = invoices_qs.filter(customer_id=customer_id)
        if staff_id and user.role == 'owner':
            invoices_qs = invoices_qs.filter(
                models.Q(created_by_id=staff_id) | models.Q(work_completed_by_id=staff_id)
            )

        invoices_list = list(invoices_qs.select_related('customer', 'car', 'created_by', 'work_completed_by'))
        total_amount = sum(inv.total for inv in invoices_list)
        my_invoices_today = [inv for inv in invoices_list if inv.work_completed_by_id == user.id or inv.created_by_id == user.id]
        my_amount = sum(inv.total for inv in my_invoices_today)
        my_count = len(my_invoices_today)

        from users.models import User as UserModel
        if user.role == 'staff':
            draft_invoices = Invoice.objects.filter(
                status='draft'
            ).filter(
                models.Q(created_by=user) | models.Q(work_completed_by=user)
            ).select_related('customer', 'car').order_by('-created_at')
        else:
            draft_invoices = Invoice.objects.filter(status='draft').select_related('customer', 'car').order_by('-created_at')

        context = {
            'total_invoices': len(invoices_list),
            'total_customers': Customer.objects.count(),
            'total_amount': total_amount,
            'my_amount': my_amount,
            'my_count': my_count,
            'recent_invoices': invoices_list[:20],
            'draft_invoices': draft_invoices,
            'draft_count': draft_invoices.count(),
            'period': period,
            'date_from': date_from,
            'date_to': date_to,
            'selected_customer': customer_id,
            'selected_staff': staff_id,
            'customers': Customer.objects.all(),
            'staff_list': UserModel.objects.all(),
        }
        return render(request, 'invoices/dashboard.html', context)


@method_decorator(login_required, name='dispatch')
class InvoiceListView(View):
    def get(self, request):
        from datetime import date, timedelta
        user = request.user
        q           = request.GET.get('q', '')
        customer_id = request.GET.get('customer')
        staff_id    = request.GET.get('staff')
        month       = request.GET.get('month')
        year        = request.GET.get('year')
        vin         = request.GET.get('vin')
        stock       = request.GET.get('stock', '')
        # Поиск по стоку/VIN — общая функция, ищет по всей базе даже для staff,
        # т.к. номер машины не привязан к конкретному сотруднику.
        searching_by_vehicle = bool(stock) or bool(vin)
        if user.role == 'staff' and not searching_by_vehicle:
            invoices = Invoice.objects.filter(
                models.Q(created_by=user) | models.Q(work_completed_by=user)
            ).distinct().select_related('customer', 'car', 'created_by', 'work_completed_by')
        else:
            invoices = Invoice.objects.all().select_related('customer', 'car', 'created_by', 'work_completed_by')
        status      = request.GET.get('status')
        sent        = request.GET.get('sent', '')
        period      = request.GET.get('period', '')
        specific_date = request.GET.get('date', '')

        if q:
            from django.db.models import Q as _Q
            invoices = invoices.filter(
                _Q(invoice_number__icontains=q) |
                _Q(customer_name__icontains=q) |
                _Q(customer__name__icontains=q) |
                _Q(car__stock_number__icontains=q) |
                _Q(car__vin__icontains=q) |
                _Q(car_info__icontains=q)
            ).distinct()
        if customer_id and customer_id.isdigit():
            invoices = invoices.filter(customer_id=customer_id)
        if staff_id and staff_id.isdigit() and user.role != 'staff':
            invoices = invoices.filter(work_completed_by_id=staff_id)
        if month:
            invoices = invoices.filter(invoice_date__month=month)
        if year:
            invoices = invoices.filter(invoice_date__year=year)
        if vin:
            invoices = invoices.filter(car__vin__icontains=vin)
        if stock:
            invoices = invoices.filter(car__stock_number__icontains=stock)
        if status:
            invoices = invoices.filter(status=status)
        if sent == '1':
            invoices = invoices.filter(email_sent=True)
        elif sent == '0':
            invoices = invoices.filter(email_sent=False)

        today = date.today()
        active_period = ''
        current_date = ''

        if period == 'today':
            invoices = invoices.filter(invoice_date=today)
            active_period = 'today'
        elif period == 'week':
            week_start = today - timedelta(days=today.weekday())
            invoices = invoices.filter(invoice_date__gte=week_start, invoice_date__lte=today)
            active_period = 'week'
        elif period == 'month':
            invoices = invoices.filter(invoice_date__year=today.year, invoice_date__month=today.month)
            active_period = 'month'
        elif specific_date:
            invoices = invoices.filter(invoice_date=specific_date)
            current_date = specific_date

        invoices = invoices.order_by('-invoice_date', '-invoice_number')

        from users.models import User
        from django.core.paginator import Paginator
        customers = Customer.objects.all()
        staff_list = User.objects.all()

        paginator = Paginator(invoices, 25)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        return render(request, 'invoices/invoice_list.html', {
            'invoices': page_obj,
            'page_obj': page_obj,
            'customers': customers,
            'staff_list': staff_list,
            'active_period': active_period,
            'current_date': current_date,
        })


@method_decorator(login_required, name='dispatch')
class InvoiceCreateView(View):
    def get(self, request):
        form = InvoiceForm(user=request.user)
        formset = InvoiceItemFormSet()
        return render(request, 'invoices/invoice_form.html', {
            'form': form,
            'formset': formset,
            'title': 'Create Invoice',
        })

    def post(self, request):
        form = InvoiceForm(request.POST, user=request.user)
        formset = InvoiceItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            invoice = form.save(commit=False)
            invoice.created_by = request.user
            if not invoice.work_completed_by:
                invoice.work_completed_by = request.user
            invoice.status = 'final'
            if request.user.role == 'owner':
                num = form.cleaned_data.get('invoice_number_override', '').strip()
                if num:
                    if Invoice.objects.filter(invoice_number=num).exists():
                        messages.error(request, f'Invoice #{num} already exists. Choose a different number.')
                        return render(request, 'invoices/invoice_form.html', {'form': form, 'formset': formset, 'title': 'Create Invoice'})
                    invoice.invoice_number = num
            invoice = get_or_create_customer_and_car(form, invoice)
            invoice.save()
            formset.instance = invoice
            formset.save()
            Invoice.objects.filter(created_by=request.user, status='draft').exclude(pk=invoice.pk).delete()
            logger.info(f'CREATED invoice={invoice.invoice_number} customer="{invoice.get_customer_display()}" total={invoice.total} user={request.user.email} status={invoice.status}')
            messages.success(request, f'Invoice #{invoice.invoice_number} created!')
            return redirect('invoice_detail', pk=invoice.pk)
        return render(request, 'invoices/invoice_form.html', {
            'form': form,
            'formset': formset,
            'title': 'Create Invoice',
        })


@method_decorator(login_required, name='dispatch')
class InvoiceEditView(View):
    def get(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        if request.user.role == 'staff' and invoice.created_by != request.user and invoice.work_completed_by != request.user:
            messages.error(request, 'Access denied')
            return redirect('invoice_list')
        initial = {
            'customer_name': invoice.get_customer_display(),
            'invoice_number_override': invoice.invoice_number,
        }
        if invoice.car:
            initial['car_make'] = invoice.car.make
            initial['car_model'] = invoice.car.model
            initial['car_vin'] = invoice.car.vin
            initial['car_stock'] = invoice.car.stock_number
        elif invoice.car_info:
            initial['car_make'] = invoice.car_info
        form = InvoiceForm(instance=invoice, user=request.user, initial=initial)
        formset = InvoiceItemFormSet(instance=invoice)
        return render(request, 'invoices/invoice_form.html', {
            'form': form,
            'formset': formset,
            'title': 'Edit Invoice',
            'invoice': invoice,
        })

    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        if request.user.role == 'staff' and invoice.created_by != request.user and invoice.work_completed_by != request.user:
            messages.error(request, 'Access denied')
            return redirect('invoice_list')
        form = InvoiceForm(request.POST, instance=invoice, user=request.user)
        formset = InvoiceItemFormSet(request.POST, instance=invoice)
        if form.is_valid() and formset.is_valid():
            invoice = form.save(commit=False)
            if request.user.role == 'owner':
                num = form.cleaned_data.get('invoice_number_override', '').strip()
                if num:
                    invoice.invoice_number = num
            invoice = get_or_create_customer_and_car(form, invoice)
            if invoice.status == 'draft':
                invoice.status = 'final'
            invoice.save()
            formset.save()
            logger.info(f'UPDATED invoice={invoice.invoice_number} customer="{invoice.get_customer_display()}" total={invoice.total} user={request.user.email} status={invoice.status}')
            messages.success(request, f'Invoice #{invoice.invoice_number} updated!')
            return redirect('invoice_detail', pk=invoice.pk)
        return render(request, 'invoices/invoice_form.html', {
            'form': form,
            'formset': formset,
            'title': 'Edit Invoice',
            'invoice': invoice,
        })


@method_decorator(login_required, name='dispatch')
class InvoiceDetailView(View):
    def get(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        return render(request, 'invoices/invoice_detail.html', {'invoice': invoice})


@method_decorator(login_required, name='dispatch')
class InvoiceActionView(View):
    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        action = request.POST.get('action')

        if action == 'send':
            emails = request.POST.getlist('emails')
            if not emails:
                if invoice.customer:
                    emails = invoice.customer.get_all_emails()
            if not emails:
                messages.error(request, 'No email address found')
                return redirect('invoice_detail', pk=pk)

            buffer = io.BytesIO()
            generate_invoice_pdf(buffer, invoice)
            buffer.seek(0)

            customer_name = invoice.get_customer_display()
            email_msg = EmailMessage(
                subject=f'Your Invoice from AutoProTinting — #{invoice.invoice_number}',
                body=(
                    f'Dear {customer_name},\n\n'
                    f'Here\'s your invoice! We appreciate your prompt payment. Thanks for your business!\n\n'
                    f'To view and print the attached invoice, double-click on the invoice icon, '
                    f'and then choose File, Print when the invoice is displayed. '
                    f'To save the invoice, copy it from this e-mail to another folder on your computer.\n\n'
                    f'If you have any questions regarding this invoice, please contact Accounting at (647) 771-1112\n\n'
                    f'Regards,\n'
                    f'Accounting Department, Autoprotinting'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=emails,
            )
            email_msg.attach(
                f'invoice_{invoice.invoice_number}.pdf',
                buffer.read(),
                'application/pdf'
            )
            def send_email():
                try:
                    email_msg.send()
                except Exception:
                    pass

            threading.Thread(target=send_email, daemon=True).start()
            invoice.status = 'final'
            invoice.email_sent = True
            invoice.save()
            messages.success(request, f'Invoice sent to {", ".join(emails)}')

        elif action == 'cancel':
            invoice.status = 'cancelled'
            invoice.save()
            messages.success(request, 'Invoice cancelled')

        elif action == 'final' or action == 'finish':
            invoice.status = 'final'
            invoice.save()
            messages.success(request, 'Invoice marked as final')

        return redirect('invoice_detail', pk=pk)


@method_decorator(login_required, name='dispatch')
class InvoicePDFView(View):
    def get(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'
        generate_invoice_pdf(response, invoice)
        return response


@method_decorator(login_required, name='dispatch')
class InvoiceDeleteView(View):
    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        if request.user.role == 'staff':
            messages.error(request, 'Access denied — staff cannot delete invoices')
            return redirect('invoice_list')
        invoice.delete()
        messages.success(request, 'Invoice deleted')
        return redirect('invoice_list')


@method_decorator(login_required, name='dispatch')
class CustomerSearchView(View):
    def get(self, request):
        q = request.GET.get('q', '')
        if len(q) >= 1:
            customers = Customer.objects.filter(name__icontains=q)[:10]
            data = [{
                'id': c.id,
                'name': c.name,
                'company': c.company or '',
                'phone': c.phone or '',
                'cars': [{'id': car.id, 'display': car.get_display_with_details(),
                          'make': car.make, 'model': car.model,
                          'vin': car.vin,
                          'stock': car.stock_number}
                         for car in c.cars.all()[:5]]
            } for c in customers]
        else:
            data = []
        return JsonResponse({'results': data})


@method_decorator(login_required, name='dispatch')
class MonthlyReportView(View):
    def get(self, request):
        invoices = Invoice.objects.exclude(status__in=['cancelled', 'draft']).select_related('customer', 'car', 'created_by')

        customer_id = request.GET.get('customer')
        staff_id = request.GET.get('staff')
        month = request.GET.get('month')
        year = request.GET.get('year', '')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        vin = request.GET.get('vin')
        stock = request.GET.get('stock')

        if customer_id:
            invoices = invoices.filter(customer_id=customer_id)
        if staff_id:
            invoices = invoices.filter(created_by_id=staff_id)
        if month:
            try:
                invoices = invoices.filter(invoice_date__month=int(month))
            except (ValueError, TypeError):
                pass
        if year and not date_from:
            try:
                invoices = invoices.filter(invoice_date__year=int(year))
            except (ValueError, TypeError):
                pass
        if date_from:
            invoices = invoices.filter(invoice_date__gte=date_from)
        if date_to:
            invoices = invoices.filter(invoice_date__lte=date_to)
        if vin:
            invoices = invoices.filter(car__vin__icontains=vin)
        if stock:
            invoices = invoices.filter(car__stock_number__icontains=stock)

        from users.models import User
        from django.core.paginator import Paginator
        from datetime import date as _mindate
        customers = Customer.objects.all()
        staff_list = User.objects.all()

        sort = request.GET.get('sort', 'date_desc')
        invoices_list = list(invoices.prefetch_related('items'))
        grand_total = sum(inv.total for inv in invoices_list)

        if sort == 'date_asc':
            invoices_list.sort(key=lambda i: i.invoice_date or _mindate.min)
        elif sort == 'total_desc':
            invoices_list.sort(key=lambda i: i.total, reverse=True)
        elif sort == 'total_asc':
            invoices_list.sort(key=lambda i: i.total)
        elif sort == 'customer':
            invoices_list.sort(key=lambda i: (i.get_customer_display() or '').upper())
        elif sort == 'number':
            invoices_list.sort(key=lambda i: i.invoice_number or '')
        else:
            invoices_list.sort(key=lambda i: i.invoice_date or _mindate.min, reverse=True)

        paginator = Paginator(invoices_list, 25)
        page_obj = paginator.get_page(request.GET.get('page'))

        qs = request.GET.copy()
        qs.pop('page', None)

        return render(request, 'invoices/monthly_report.html', {
            'invoices': page_obj,
            'page_obj': page_obj,
            'customers': customers,
            'staff_list': staff_list,
            'total': grand_total,
            'month': month,
            'year': year,
            'sort': sort,
            'querystring': qs.urlencode(),
        })


@method_decorator(login_required, name='dispatch')
class MonthlyReportPDFView(View):
    def get(self, request):
        invoices = Invoice.objects.exclude(status__in=['cancelled', 'draft']).select_related('customer', 'car', 'created_by')

        customer_id = request.GET.get('customer')
        staff_id = request.GET.get('staff')
        month = request.GET.get('month')
        year = request.GET.get('year')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        vin = request.GET.get('vin')

        if customer_id:
            invoices = invoices.filter(customer_id=customer_id)
        if staff_id:
            invoices = invoices.filter(created_by_id=staff_id)
        if month:
            try:
                invoices = invoices.filter(invoice_date__month=int(month))
            except (ValueError, TypeError):
                pass
        if year and not date_from:
            try:
                invoices = invoices.filter(invoice_date__year=int(year))
            except (ValueError, TypeError):
                pass
        if date_from:
            invoices = invoices.filter(invoice_date__gte=date_from)
        if date_to:
            invoices = invoices.filter(invoice_date__lte=date_to)
        if vin:
            invoices = invoices.filter(car__vin__icontains=vin)
        if stock:
            invoices = invoices.filter(car__stock_number__icontains=stock)

        customer = None
        if customer_id:
            from customers.models import Customer as C
            customer = C.objects.filter(pk=customer_id).first()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Invoice_Statement.pdf"'
        generate_monthly_report_pdf(response, invoices, customer=customer)
        return response


@method_decorator(login_required, name='dispatch')
class SendStatementView(View):
    def post(self, request):
        customer_id = request.POST.get('customer_id')
        customer = get_object_or_404(Customer, pk=customer_id)
        emails = customer.get_all_emails()

        if not emails:
            messages.error(request, 'No email address found for this customer')
            return redirect('monthly_report')

        invoices = Invoice.objects.filter(customer=customer).exclude(status__in=['cancelled', 'draft'])
        month = request.POST.get('month')
        year = request.POST.get('year')
        if month:
            invoices = invoices.filter(invoice_date__month=month)
        if year:
            invoices = invoices.filter(invoice_date__year=year)

        buffer = io.BytesIO()
        generate_monthly_report_pdf(buffer, invoices, customer=customer)
        buffer.seek(0)

        email_msg = EmailMessage(
            subject='Statement from AUTOPROTINTING INC',
            body=(
                f'Dear {customer.name},\n\n'
                f'Your statement is attached. Please remit payment at your earliest convenience.\n'
                f'Thank you for your business - we appreciate it very much.\n\n'
                f'Have a great day!\n'
                f'AUTOPROTINTING INC'
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=emails,
        )
        email_msg.attach('statement.pdf', buffer.read(), 'application/pdf')
        def send_statement():
            try:
                email_msg.send()
            except Exception:
                pass

        threading.Thread(target=send_statement, daemon=True).start()
        messages.success(request, f'Statement sent to {", ".join(emails)}')
        return redirect('monthly_report')

@method_decorator(login_required, name='dispatch')
class StatementPreviewView(View):
    def get(self, request):
        customer_id = request.GET.get('customer')
        from customers.models import Customer as C
        customer = C.objects.filter(pk=customer_id).first() if customer_id else None
        invoices = Invoice.objects.filter(customer=customer).exclude(status__in=['cancelled', 'draft']) if customer else Invoice.objects.exclude(status__in=['cancelled', 'draft'])[:5]
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="preview.pdf"'
        generate_monthly_report_pdf(response, invoices, customer=customer)
        return response


@method_decorator(login_required, name='dispatch')
class InvoiceHistoryReportView(View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        invoices = Invoice.objects.all().select_related('customer', 'car', 'created_by', 'work_completed_by')

        customer_id = request.GET.get('customer')
        staff_id = request.GET.get('staff')
        month = request.GET.get('month')
        year = request.GET.get('year')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        vin = request.GET.get('vin')

        if customer_id:
            invoices = invoices.filter(customer_id=customer_id)
        if staff_id:
            invoices = invoices.filter(created_by_id=staff_id)
        if month:
            try:
                invoices = invoices.filter(invoice_date__month=int(month))
            except (ValueError, TypeError):
                pass
        if year and not date_from:
            try:
                invoices = invoices.filter(invoice_date__year=int(year))
            except (ValueError, TypeError):
                pass
        if date_from:
            invoices = invoices.filter(invoice_date__gte=date_from)
        if date_to:
            invoices = invoices.filter(invoice_date__lte=date_to)
        if vin:
            invoices = invoices.filter(car__vin__icontains=vin)
        if stock:
            invoices = invoices.filter(car__stock_number__icontains=stock)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Invoice History'

        headers = ['Invoice Number', 'Invoice Date', 'Total', 'Customer', 'Description',
                   'Asset Name', 'Stock #', 'VIN Number', 'User', 'Work Completed By']

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row = 2
        for inv in invoices:
            items = list(inv.items.all())
            if not items:
                items = [None]
            for i, item in enumerate(items):
                ws.cell(row=row, column=1, value=inv.invoice_number if i == 0 else '')
                ws.cell(row=row, column=2, value=inv.invoice_date.strftime('%m/%d/%Y') if i == 0 else '')
                ws.cell(row=row, column=3, value=float(inv.total) if i == 0 else '')
                ws.cell(row=row, column=4, value=inv.get_customer_display() if i == 0 else '')
                ws.cell(row=row, column=5, value=item.description if item else '')
                asset_name = f'{inv.car.make} {inv.car.model}' if inv.car and i == 0 else (inv.car_info if i == 0 else '')
                ws.cell(row=row, column=6, value=asset_name)
                ws.cell(row=row, column=7, value=inv.car.stock_number if inv.car and inv.car.stock_number and i == 0 else '')
                ws.cell(row=row, column=8, value=inv.car.vin if inv.car and inv.car.vin and i == 0 else '')
                ws.cell(row=row, column=9, value=inv.created_by.get_full_name() if inv.created_by and i == 0 else '')
                ws.cell(row=row, column=10, value=inv.work_completed_by.get_full_name() if inv.work_completed_by and i == 0 else '')
                row += 1

        from openpyxl.styles import Border, Side
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        fixed_widths = [16, 12, 10, 25, 35, 18, 12, 16, 20, 20]
        for i, col in enumerate(ws.columns):
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = fixed_widths[i] if i < len(fixed_widths) else 15
            for cell in col:
                cell.border = border
                if cell.row > 1:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Invoice_History_Report.xlsx"'
        response.write(buffer.read())
        return response


@method_decorator(login_required, name='dispatch')
@method_decorator(login_required, name='dispatch')
class MobileReportsView(View):
    def get(self, request):
        if request.user.role not in ('owner', 'accountant'):
            return redirect('dashboard')
        return render(request, 'invoices/mobile_reports.html')


class CrewReportView(View):
    def get(self, request):
        if request.user.role != 'owner':
            return redirect('dashboard')
        from users.models import User
        from django.utils.dateparse import parse_date

        start = request.GET.get('start')
        end = request.GET.get('end')
        month = request.GET.get('month')
        year = request.GET.get('year')
        # если выбран месяц без года — дефолт на текущий год
        if month and not year:
            from datetime import date as _date
            year = str(_date.today().year)

        invoices = Invoice.objects.filter(status='final').select_related('work_completed_by')
        if start:
            invoices = invoices.filter(invoice_date__gte=parse_date(start))
        if end:
            invoices = invoices.filter(invoice_date__lte=parse_date(end))
        if month:
            try:
                invoices = invoices.filter(invoice_date__month=int(month))
            except (ValueError, TypeError):
                pass
        if year:
            try:
                invoices = invoices.filter(invoice_date__year=int(year))
            except (ValueError, TypeError):
                pass

        staff_users = User.objects.filter(role='staff')
        rows = []
        for user in staff_users:
            user_invoices = [inv for inv in invoices if inv.work_completed_by_id == user.id]
            count = len(user_invoices)
            subtotal = sum(inv.subtotal for inv in user_invoices)
            gross = sum(inv.total for inv in user_invoices)
            rows.append({'user': user, 'count': count, 'subtotal': subtotal, 'gross': gross})

        return render(request, 'invoices/crew_report.html', {
            'rows': rows,
            'total_count': sum(r['count'] for r in rows),
            'total_subtotal': sum(r['subtotal'] for r in rows),
            'total_gross': sum(r['gross'] for r in rows),
            'start': start or '',
            'end': end or '',
            'month': month or '',
            'year': year or '',
        })


@method_decorator(login_required, name='dispatch')
class CustomerReportView(View):
    def get(self, request):
        if request.user.role != 'owner':
            return redirect('dashboard')
        from django.utils.dateparse import parse_date
        from django.core.paginator import Paginator
        from datetime import date as _date
        start = request.GET.get('start')
        end = request.GET.get('end')
        month = request.GET.get('month')
        year = request.GET.get('year')
        sort = request.GET.get('sort', 'gross_desc')
        if month and not year:
            year = str(_date.today().year)
        invoices = Invoice.objects.exclude(status__in=['cancelled', 'draft']).select_related('customer').prefetch_related('items')
        if start:
            invoices = invoices.filter(invoice_date__gte=parse_date(start))
        if end:
            invoices = invoices.filter(invoice_date__lte=parse_date(end))
        if month:
            try:
                invoices = invoices.filter(invoice_date__month=int(month))
            except (ValueError, TypeError):
                pass
        if year:
            try:
                invoices = invoices.filter(invoice_date__year=int(year))
            except (ValueError, TypeError):
                pass
        invoices_list = list(invoices)
        customer_map = {}
        for inv in invoices_list:
            key = inv.customer_id if inv.customer_id else f'_noname_{inv.customer_name}'
            if key not in customer_map:
                customer_map[key] = {
                    'name': inv.get_customer_display(),
                    'count': 0, 'subtotal': 0, 'gross': 0
                }
            customer_map[key]['count'] += 1
            customer_map[key]['subtotal'] += inv.subtotal
            customer_map[key]['gross'] += inv.total
        rows = list(customer_map.values())
        if sort == 'gross_asc':
            rows.sort(key=lambda x: x['gross'])
        elif sort == 'count_desc':
            rows.sort(key=lambda x: x['count'], reverse=True)
        elif sort == 'count_asc':
            rows.sort(key=lambda x: x['count'])
        elif sort == 'name_asc':
            rows.sort(key=lambda x: (x['name'] or '').upper())
        elif sort == 'name_desc':
            rows.sort(key=lambda x: (x['name'] or '').upper(), reverse=True)
        else:
            rows.sort(key=lambda x: x['gross'], reverse=True)
        total_count = sum(r['count'] for r in rows)
        total_subtotal = sum(r['subtotal'] for r in rows)
        total_gross = sum(r['gross'] for r in rows)
        paginator = Paginator(rows, 100)
        page_obj = paginator.get_page(request.GET.get('page'))
        qs = request.GET.copy()
        qs.pop('page', None)
        return render(request, 'invoices/customer_report.html', {
            'rows': page_obj,
            'page_obj': page_obj,
            'total_count': total_count,
            'total_subtotal': total_subtotal,
            'total_gross': total_gross,
            'start': start or '',
            'end': end or '',
            'month': month or '',
            'year': year or '',
            'sort': sort,
            'querystring': qs.urlencode(),
        })


@method_decorator(csrf_exempt, name='dispatch')
class SaveDraftView(View):
    def post(self, request):
        import json
        from django.http import JsonResponse
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'unauthorized'}, status=401)
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'invalid json'}, status=400)

        customer_name = data.get('customer_name', '').strip()
        car_make = data.get('car_make', '').strip()
        if not customer_name and not car_make:
            return JsonResponse({'error': 'empty'}, status=400)

        # get or create customer
        from customers.models import Customer
        customer = None
        if customer_name:
            customer = Customer.objects.filter(name__iexact=customer_name).first()
            if not customer:
                customer = Customer.objects.create(name=customer_name.upper())

        # get or create car
        from cars.models import Car
        car = None
        car_vin = data.get('car_vin', '').strip().upper()
        car_model = data.get('car_model', '').strip().upper()
        car_stock = data.get('car_stock', '').strip().upper()
        if customer and car_make:
            if car_vin:
                car = Car.objects.filter(vin=car_vin).first()
            if not car:
                car = Car.objects.create(
                    customer=customer,
                    make=car_make.upper(),
                    model=car_model,
                    vin=car_vin,
                    stock_number=car_stock,
                )

        from datetime import date
        from dateutil.parser import parse as parse_date_str
        invoice_date_str = data.get('invoice_date') or str(date.today())
        try:
            invoice_date = parse_date_str(invoice_date_str).date()
        except Exception:
            invoice_date = date.today()

        # delete old drafts by this user without invoice_number
        Invoice.objects.filter(
            created_by=request.user,
            status='draft',
        ).delete()

        invoice = Invoice.objects.create(
            customer=customer,
            car=car,
            status='draft',
            invoice_date=invoice_date,
            po_number=data.get('po_number', ''),
            notes=data.get('notes', ''),
            created_by=request.user,
            work_completed_by=request.user,
        )

        services = data.get('services', [])
        from invoices.models import InvoiceItem
        for s in services:
            desc = s.get('desc', '').strip()
            price = s.get('price', 0)
            if desc:
                try:
                    price = float(price)
                except Exception:
                    price = 0
                InvoiceItem.objects.create(invoice=invoice, description=desc.upper(), price=price)

        return JsonResponse({'ok': True, 'pk': invoice.pk})


@method_decorator(csrf_exempt, name='dispatch')
class DeleteDraftView(View):
    def post(self, request):
        from django.http import JsonResponse
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'unauthorized'}, status=401)
        Invoice.objects.filter(created_by=request.user, status='draft').delete()
        return JsonResponse({'ok': True})


class DownloadDatabaseView(View):
    def get(self, request):
        if request.user.role != 'owner':
            return redirect('dashboard')
        import subprocess
        from django.conf import settings as django_settings
        from datetime import date as _date
        dbcfg = django_settings.DATABASES['default']
        db_name = dbcfg['NAME']
        site_name = 'devapt' if 'dev' in str(db_name) else 'apt'
        if 'postgresql' in dbcfg['ENGINE']:
            dump = subprocess.run(
                ['docker', 'exec', 'apt_postgres', 'pg_dump', '-U', 'postgres', str(db_name)],
                capture_output=True
            )
            if dump.returncode != 0:
                return HttpResponse('Backup failed: ' + dump.stderr.decode()[:500], status=500)
            response = HttpResponse(dump.stdout, content_type='application/sql')
            response['Content-Disposition'] = f'attachment; filename="db_{site_name}_{_date.today()}.sql"'
            return response
        else:
            with open(db_name, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="db_{site_name}_{_date.today()}.sqlite3"'
            return response
