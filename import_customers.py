import os
import sys
import django

# Usage: python3 import_customers.py <sheet_name>
# sheet_name: 'APTTract costomers' or 'DEVAPT costomers.'

sheet_name = sys.argv[1] if len(sys.argv) > 1 else 'APTTract costomers'

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'apt_tract.settings')
django.setup()

from openpyxl import load_workbook
from customers.models import Customer, CustomerEmail

wb = load_workbook('/tmp/customers.xlsx', read_only=True)
ws = wb[sheet_name]

rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data_rows = rows[1:]

created = 0
skipped = 0

for row in data_rows:
    if not row[0]:
        continue

    name = str(row[0]).strip()
    address = str(row[2]).strip() if row[2] else ''
    postal_code = str(row[3]).strip() if row[3] else ''
    city = str(row[4]).strip() if row[4] else ''
    province = str(row[5]).strip() if row[5] else ''
    country = str(row[6]).strip() if row[6] else 'Canada'
    emails_raw = str(row[8]).strip() if row[8] else ''
    phone = str(row[9]).strip() if row[9] else ''

    # Skip if already exists
    if Customer.objects.filter(name__iexact=name).exists():
        print(f'SKIP (exists): {name}')
        skipped += 1
        continue

    customer = Customer.objects.create(
        name=name,
        address=address,
        postal_code=postal_code,
        city=city,
        province=province,
        country=country if country else 'Canada',
        phone=phone,
    )

    # Parse multiple emails
    if emails_raw and emails_raw != 'None':
        emails = [e.strip() for e in emails_raw.replace(';', ',').split(',') if e.strip()]
        for i, email in enumerate(emails):
            if '@' in email:
                CustomerEmail.objects.create(
                    customer=customer,
                    email=email,
                    is_primary=(i == 0)
                )

    print(f'CREATED: {name} | emails: {len([e for e in emails_raw.split(",") if "@" in e])}')
    created += 1

print(f'\nDone! Created: {created}, Skipped: {skipped}')
